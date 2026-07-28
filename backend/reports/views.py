from datetime import date, timedelta

from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.http import JsonResponse
from django.shortcuts import render

from income.models import Income
from expenses.models import Expense
from production.models import Production
from raw_materials.models import MaterialPurchase


def _month_start(d):
    return d.replace(day=1)


def _last_six_months():
    cursor = _month_start(date.today())
    months = []
    for _ in range(6):
        months.append(cursor)
        cursor = _month_start(cursor - timedelta(days=1))
    return list(reversed(months))


def reports(request):
    revenue = Income.objects.aggregate(t=Sum("amount"))["t"] or 0
    expenses_total = Expense.objects.aggregate(t=Sum("amount"))["t"] or 0
    material_cost = MaterialPurchase.objects.aggregate(t=Sum("total_cost"))["t"] or 0

    total_cost = expenses_total + material_cost
    net_profit = revenue - total_cost
    margin = round((net_profit / revenue) * 100) if revenue else 0

    runs = Production.objects.all()
    total_produced = runs.aggregate(t=Sum("produced_quantity"))["t"] or 0

    # monthly table
    months = _last_six_months()

    def monthly(model, field="amount"):
        rows = (model.objects
                .filter(date__gte=months[0])
                .annotate(m=TruncMonth("date"))
                .values("m")
                .annotate(total=Sum(field)))
        return {r["m"].strftime("%Y-%m"): r["total"] for r in rows}

    inc_by_month = monthly(Income)
    exp_by_month = monthly(Expense)
    prod_by_month = monthly(Production, "produced_quantity")

    table = []
    for m in months:
        key = m.strftime("%Y-%m")
        inc = inc_by_month.get(key, 0) or 0
        exp = exp_by_month.get(key, 0) or 0
        profit = inc - exp
        table.append({
            "month": m.strftime("%B %Y"),
            "revenue": inc,
            "expenses": exp,
            "profit": profit,
            "production": prod_by_month.get(key, 0) or 0,
            "performance": "good" if profit > 0 else ("flat" if profit == 0 else "poor"),
        })

    return render(request, "reports.html", {
        "revenue": revenue,
        "expenses_total": expenses_total,
        "material_cost": material_cost,
        "total_cost": total_cost,
        "net_profit": net_profit,
        "margin": margin,
        "total_produced": total_produced,
        "completed_runs": runs.filter(status="completed").count(),
        "pending_runs": runs.exclude(status="completed").count(),
        "monthly_table": table,
    })


def reports_charts(request):
    months = _last_six_months()

    def monthly(model):
        rows = (model.objects
                .filter(date__gte=months[0])
                .annotate(m=TruncMonth("date"))
                .values("m")
                .annotate(total=Sum("amount")))
        return {r["m"].strftime("%Y-%m"): float(r["total"]) for r in rows}

    inc = monthly(Income)
    exp = monthly(Expense)
    keys = [m.strftime("%Y-%m") for m in months]

    return JsonResponse({
        "labels": [m.strftime("%b") for m in months],
        "revenue": [inc.get(k, 0) for k in keys],
        "expenses": [exp.get(k, 0) for k in keys],
    })

def reports_export(request):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    money = '"NPR "#,##0'

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def autosize(ws):
        for col in ws.columns:
            width = max(len(str(c.value or "")) for c in col) + 4
            ws.column_dimensions[col[0].column_letter].width = width

    # --- Sheet 1: Monthly summary ---
    ws = wb.active
    ws.title = "Monthly Summary"
    ws.append(["Month", "Revenue", "Expenses", "Net Profit", "Production"])

    months = _last_six_months()

    def monthly(model, field="amount"):
        rows = (model.objects
                .filter(date__gte=months[0])
                .annotate(m=TruncMonth("date"))
                .values("m")
                .annotate(total=Sum(field)))
        return {r["m"].strftime("%Y-%m"): r["total"] for r in rows}

    inc_by_month = monthly(Income)
    exp_by_month = monthly(Expense)
    prod_by_month = monthly(Production, "produced_quantity")

    for m in months:
        key = m.strftime("%Y-%m")
        inc = inc_by_month.get(key, 0) or 0
        exp = exp_by_month.get(key, 0) or 0
        ws.append([
            m.strftime("%B %Y"),
            float(inc),
            float(exp),
            float(inc - exp),
            prod_by_month.get(key, 0) or 0,
        ])

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = money

    style_header(ws)
    autosize(ws)

    # --- Sheet 2: Sales ---
    ws2 = wb.create_sheet("Sales")
    ws2.append(["Date", "Invoice", "Customer", "Product",
                "Pairs", "Unit Price", "Amount", "Payment"])

    for s in Income.objects.select_related("customer", "product"):
        ws2.append([
            s.date, s.invoice_no, s.customer.name, s.product.name,
            s.pairs_sold, float(s.unit_price), float(s.amount),
            s.get_payment_status_display(),
        ])

    for row in ws2.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = money

    style_header(ws2)
    autosize(ws2)

    # --- Sheet 3: Expenses ---
    ws3 = wb.create_sheet("Expenses")
    ws3.append(["Date", "Category", "Description", "Amount"])

    for e in Expense.objects.select_related("category"):
        ws3.append([e.date, e.category.name, e.description, float(e.amount)])

    for row in ws3.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = money

    style_header(ws3)
    autosize(ws3)

    # --- Sheet 4: Production ---
    ws4 = wb.create_sheet("Production")
    ws4.append(["Date", "Batch", "Product", "Supervisor",
                "Target", "Produced", "Defective", "Status"])

    for r in Production.objects.select_related("product"):
        ws4.append([
            r.date, r.batch_no, r.product.name, r.supervisor,
            r.target_quantity, r.produced_quantity,
            r.defective_quantity, r.get_status_display(),
        ])

    style_header(ws4)
    autosize(ws4)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"FactoryFlow-Report-{date.today():%Y-%m-%d}.xlsx"

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response